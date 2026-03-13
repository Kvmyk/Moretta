"""
PrivateProxy — XLSX file parser.
Extracts text content from Microsoft Excel spreadsheets.
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger("privateproxy.parsers.xlsx")


from openpyxl.utils import get_column_letter

def parse_xlsx(file_path: Path) -> dict[str, Any]:
    """
    Extract text and structure from an XLSX file.
    """
    wb = load_workbook(str(file_path), read_only=True, data_only=True)
    text_parts: list[str] = []
    preview_sheets: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        text_parts.append(f"[Arkusz: {sheet_name}]")
        
        rows_data = []
        preview_row_count = 0
        
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            row_cells = []
            has_value = False
            for col_idx, cell_value in enumerate(row, start=1):
                val = str(cell_value).strip() if cell_value is not None else ""
                if val:
                    col_letter = get_column_letter(col_idx)
                    text_parts.append(f"{col_letter}{row_idx}: {val}")
                    has_value = True
                row_cells.append(val)
            
            if has_value:
                if preview_row_count < 50:
                    rows_data.append(row_cells)
                    preview_row_count += 1
        
        preview_sheets.append({
            "name": sheet_name,
            "rows": rows_data
        })

    wb.close()

    full_text = "\n".join(text_parts)
    logger.info(
        f"Parsed XLSX: {file_path.name} — {len(wb.sheetnames)} sheets, {len(full_text)} chars"
    )
    return {
        "text": full_text,
        "preview_data": {
            "type": "spreadsheet",
            "sheets": preview_sheets
        }
    }
