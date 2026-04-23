import base64
import json
import sqlite3
from pathlib import Path

import pytest



def _create_legacy_store_db(db_path: Path, key: str, filename: str, created_at: str) -> None:
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS files (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL,
                created_at TEXT
            )
            """
        )
        conn.execute(
            "INSERT INTO files (key, value, created_at) VALUES (?, ?, ?)",
            (
                key,
                json.dumps({"filename": filename, "ext": Path(filename).suffix.lower()}),
                created_at,
            ),
        )
        conn.commit()


def test_migration_preserves_docx_blob_content(tmp_path: Path):
    pytest.importorskip("docx")
    pytest.importorskip("psycopg")
    from docx import Document
    from parsers.docx_parser import parse_docx
    from scripts.migrate_sqlite_to_postgres import _load_legacy_store

    key = "legacy-docx-1"
    store_db = tmp_path / "store.db"
    blob_dir = tmp_path / "blobs"
    blob_dir.mkdir()

    docx_source = tmp_path / "source.docx"
    doc = Document()
    doc.add_paragraph("Umowa: Jan Kowalski, PESEL 92010212345")
    doc.save(str(docx_source))
    original_bytes = docx_source.read_bytes()

    _create_legacy_store_db(
        db_path=store_db,
        key=key,
        filename="umowa.docx",
        created_at="2026-04-23T10:00:00+00:00",
    )
    (blob_dir / f"{key}.original_bytes").write_bytes(original_bytes)

    with sqlite3.connect(store_db) as conn:
        rows = list(_load_legacy_store(conn, "files", blob_dir, fernet=None))

    assert len(rows) == 1
    migrated_key, value_json, created_at, blob_json = rows[0]
    assert migrated_key == key
    assert json.loads(value_json)["filename"] == "umowa.docx"
    assert created_at == "2026-04-23T10:00:00+00:00"

    migrated_blob_bytes = base64.b64decode(json.loads(blob_json)["original_bytes"])
    migrated_docx = tmp_path / "migrated.docx"
    migrated_docx.write_bytes(migrated_blob_bytes)

    parsed = parse_docx(migrated_docx)
    assert parsed["preview_data"]["type"] == "document"
    assert "Jan Kowalski" in parsed["text"]


def test_migration_preserves_xlsx_blob_content(tmp_path: Path):
    pytest.importorskip("openpyxl")
    pytest.importorskip("psycopg")
    from openpyxl import Workbook
    from parsers.xlsx_parser import parse_xlsx
    from scripts.migrate_sqlite_to_postgres import _load_legacy_store

    key = "legacy-xlsx-1"
    store_db = tmp_path / "store.db"
    blob_dir = tmp_path / "blobs"
    blob_dir.mkdir()

    xlsx_source = tmp_path / "source.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Arkusz1"
    ws["A1"] = "Imie i nazwisko"
    ws["A2"] = "Jan Kowalski"
    ws["B1"] = "PESEL"
    ws["B2"] = "92010212345"
    wb.save(str(xlsx_source))
    original_bytes = xlsx_source.read_bytes()

    _create_legacy_store_db(
        db_path=store_db,
        key=key,
        filename="raport.xlsx",
        created_at="2026-04-23T10:00:00+00:00",
    )
    (blob_dir / f"{key}.original_bytes").write_bytes(original_bytes)

    with sqlite3.connect(store_db) as conn:
        rows = list(_load_legacy_store(conn, "files", blob_dir, fernet=None))

    assert len(rows) == 1
    migrated_key, value_json, created_at, blob_json = rows[0]
    assert migrated_key == key
    assert json.loads(value_json)["filename"] == "raport.xlsx"
    assert created_at == "2026-04-23T10:00:00+00:00"

    migrated_blob_bytes = base64.b64decode(json.loads(blob_json)["original_bytes"])
    migrated_xlsx = tmp_path / "migrated.xlsx"
    migrated_xlsx.write_bytes(migrated_blob_bytes)

    parsed = parse_xlsx(migrated_xlsx)
    assert parsed["preview_data"]["type"] == "spreadsheet"
    assert "Jan Kowalski" in parsed["text"]
